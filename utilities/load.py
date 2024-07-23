# pylint: skip-file
# mypy: ignore-errors

import datetime
import json
import pathlib
import typing

import openpyxl
from openpyxl.workbook.workbook import Workbook
from rich import print

from tmo import config

data: dict[str, dict[str, dict[str, typing.Any]]] = {}
max_row = 0
num_name_map = {}

num_fix_map = config.load.names

data = json.loads(pathlib.Path("bills.json").read_text())


def user_1(date: str, wb: Workbook) -> int:
    ws = wb["Charges"]
    data[date] = []
    for (
        num,
        name,
        _,
        phone,
        line,
        ins,
        usage,
        _,
        minutes,
        messages,
        gigs,
    ) in ws.iter_cols(2, 100, 1, 11, values_only=True):
        if num is None:
            break

        data[date].append(
            {
                "name": name,
                "num": num,
                "phone": phone,
                "line": line,
                "insurance": ins,
                "usage": usage,
                "minutes": minutes,
                "messages": messages,
                "data": gigs,
            }
        )


def user_2(date: str, wb: Workbook) -> int:
    ws = wb["Charges"]
    data[date] = {}
    for (
        num,
        name,
        _,
        phone,
        line,
        ins,
        usage,
    ) in ws.iter_cols(2, 100, 1, 7, values_only=True):
        if num is None:
            break

        data[date][name] = {
            "name": name,
            "num": num,
            "phone": phone,
            "line": line,
            "insurance": ins,
            "usage": usage,
        }

        if num not in num_name_map:
            num_name_map[num] = name

    for num, minutes, messages, gigs in ws.iter_rows(15, 30, 2, 5, values_only=True):
        if num is None:
            break

        for old, new in num_fix_map.values():
            if num == old:
                num = new
                break

        name = num_name_map[num]

        data[date][name].update(
            {
                "minutes": minutes,
                "messages": messages,
                "data": gigs,
            }
        )


def user_3(date: str, wb: Workbook) -> None:
    ws = wb["Charges"]

    data[date] = {}
    for (
        num,
        name,
        _,
        phone,
        line,
        ins,
        usage,
    ) in ws.iter_cols(2, 100, 1, 7, values_only=True):
        if num is None:
            break

        data[date][name] = {
            "name": name,
            "num": num,
            "phone": phone,
            "line": line,
            "insurance": ins,
            "usage": usage,
        }

        if num not in num_name_map:
            num_name_map[num] = name

    ws = wb["Statistics"]

    for num, minutes, messages, gigs in ws.iter_rows(1, 100, 1, 4, values_only=True):
        if messages is None:
            break

        if num is None:
            continue

        for old, new in num_fix_map.values():
            if num == old:
                num = new
                break

        name = num_name_map[num]

        data[date][name].update(
            {
                "minutes": minutes,
                "messages": messages,
                "data": gigs,
            }
        )


NOTES = """
- 2019.06.31 and 2019.06.27 seem to be strange...
"""

SKIP = ["2019.06.31", "2019.06.27"]
mapper = [
    (datetime.date(2020, 1, 30), user_1),
    (datetime.date(2019, 4, 18), user_2),
    (datetime.date(2019, 1, 8), user_3),
    (datetime.date(1000, 1, 1), lambda *k: k),
]


def parse(file):
    fname = file.stem.replace("Bill ", "")

    if fname in SKIP:
        return

    attempt = True
    if attempt:
        try:
            bill_date = datetime.date(*map(int, fname.split(".")))

            for date, func in mapper:
                if bill_date > date:
                    return func(fname, openpyxl.load_workbook(file, data_only=True))

        except ValueError:
            print(file)

    else:
        bill_date = datetime.date(*map(int, fname.split(".")))

        for date, func in mapper:
            if bill_date > date:
                return func(fname, openpyxl.load_workbook(file, data_only=True))


def convert():
    for index, file in enumerate(sorted(pathlib.Path("bills").glob("*.xlsx"), reverse=True)):
        if index == 3:
            break
        print(file)
        parse(file)

    pathlib.Path("bills.json").write_text(json.dumps(data, indent="\t", sort_keys=True))


if __name__ == "__main__":
    convert()
